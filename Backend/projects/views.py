# backend/projects/views.py
import openpyxl
import requests
import json
from django.http import HttpResponse
from rest_framework import viewsets, permissions, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Sum, Q
from rest_framework.permissions import AllowAny
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
import time
from django.contrib.auth import get_user_model
from django.utils import timezone

# ایمپورت مدل‌ها
from .models import (
    Project, Scenario, CalendarEvent, WeeklyReport,
    ProjectFile, ProjectPayment, ProjectExpense, Notification,
    SalaryPayment, GeneralExpense, ScenarioComment, ActivityLog,
    ChatRoom, ChatMessage,
    PaymentMethod, AgencyInfo, Package,
    ExtraService, ServiceRequest, AgencyFile, TargetAudience, Task, FileComment, StickyNote,
    WorkflowRule, TimeLog, SharedLink, DashboardConfig, Lead,
)

# ایمپورت سریالایزرها
from .serializers import (
    ProjectListSerializer, ProjectDetailSerializer, ProjectSerializer,
    ScenarioSerializer, CalendarEventSerializer, WeeklyReportSerializer,
    ProjectFileSerializer, ProjectPaymentSerializer, ProjectExpenseSerializer,
    NotificationSerializer, SalaryPaymentSerializer, GeneralExpenseSerializer,
    ScenarioCommentSerializer, ActivityLogSerializer,
    ChatRoomSerializer, ChatMessageSerializer,
    PaymentMethodSerializer, AgencyInfoSerializer, PackageSerializer,
    ExtraServiceSerializer, ServiceRequestSerializer, GlobalCalendarEventSerializer, AgencyFileSerializer,
    TargetAudienceSerializer, TaskSerializer, FileCommentSerializer, StickyNoteSerializer, WorkflowRuleSerializer,
    TimeLogSerializer, SharedLinkSerializer,
    DashboardConfigSerializer, LeadSerializer
)

User = get_user_model()

# ✅ تنظیمات OpenRouter با مدل جدید شیائومی
OPENROUTER_API_KEY = "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"
# 🔹 مدل انتخاب شده: شیائومی (رایگان و سریع)
OPENROUTER_MODEL = "xiaomi/mimo-v2-flash:free"


# --- Permissions (سطح دسترسی) ---

class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated and (
                request.user.role == 'admin' or request.user.is_superuser)


class IsOwnerOrAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated

    def has_object_permission(self, request, view, obj):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.user.role == 'admin' or request.user.is_superuser:
            return True
        if isinstance(obj, Project):
            return obj.client_user == request.user
        return False


class IsProjectTeamMember(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated

    def has_object_permission(self, request, view, obj):
        user = request.user
        if not user.is_authenticated: return False
        if user.role == 'admin' or user.is_superuser: return True

        if isinstance(obj, Project):
            return (
                    obj.client_user == user or
                    user in obj.writers.all() or
                    user in obj.videographers.all() or
                    user in obj.editors.all() or
                    user in obj.designers.all() or
                    user in obj.social_admins.all()
            )

        if hasattr(obj, 'project') and obj.project:
            return (
                    obj.project.client_user == user or
                    user in obj.project.writers.all() or
                    user in obj.project.videographers.all() or
                    user in obj.project.editors.all() or
                    user in obj.project.designers.all() or
                    user in obj.project.social_admins.all()
            )
        return False


class IsParentProjectOwnerOrAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.user.role == 'admin' or request.user.is_superuser:
            return True

        project_pk = view.kwargs.get('project_pk')
        if not project_pk:
            return False

        try:
            project = Project.objects.get(pk=project_pk)
        except Project.DoesNotExist:
            return False

        return project.client_user == request.user


# --- ViewSets ---

class ProjectViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['project_name', 'page_username']

    def get_serializer_class(self):
        if self.action == 'list': return ProjectListSerializer
        if self.action == 'retrieve': return ProjectDetailSerializer
        return ProjectSerializer

    def get_permissions(self):
        if self.action in ['create', 'destroy', 'trash', 'restore', 'hard_delete']:
            return [IsAdminUser()]
        return [permissions.IsAuthenticated(), IsProjectTeamMember()]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated: return Project.objects.none()

        base_qs = Project.objects.filter(is_deleted=False).order_by('-id')

        if user.role == 'admin' or user.is_superuser:
            return base_qs
        elif user.role == 'client':
            return base_qs.filter(client_user=user)
        else:
            return base_qs.filter(
                Q(writers=user) | Q(videographers=user) | Q(editors=user) |
                Q(designers=user) | Q(social_admins=user)
            ).distinct()

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.save()
        log_activity(self.request.user, 'حذف (موقت)', 'پروژه', f"پروژه {instance.project_name} به سطل زباله منتقل شد.",
                     instance, self.request)

    @action(detail=False, methods=['get'])
    def trash(self, request):
        deleted_projects = Project.objects.filter(is_deleted=True).order_by('-id')
        serializer = ProjectListSerializer(deleted_projects, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        try:
            project = Project.objects.get(pk=pk)
            project.is_deleted = False
            project.save()
            log_activity(request.user, 'بازگردانی', 'پروژه', f"پروژه {project.project_name} بازگردانی شد.", project,
                         request)
            return Response({'status': 'restored'})
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=404)

    @action(detail=True, methods=['delete'])
    def hard_delete(self, request, pk=None):
        try:
            project = Project.objects.get(pk=pk)
            log_activity(request.user, 'حذف دائم', 'پروژه', f"پروژه {project.project_name} برای همیشه حذف شد.", None,
                         request)
            project.delete()
            return Response({'status': 'deleted permanently'})
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=404)


class ScenarioViewSet(viewsets.ModelViewSet):
    serializer_class = ScenarioSerializer
    permission_classes = [permissions.IsAuthenticated, IsProjectTeamMember]

    def get_queryset(self):
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            return Scenario.objects.filter(project_id=project_pk)
        return Scenario.objects.all()

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs.get('project_pk'))
        serializer.save(project=project)


class CalendarEventViewSet(viewsets.ModelViewSet):
    serializer_class = CalendarEventSerializer
    permission_classes = [permissions.IsAuthenticated, IsProjectTeamMember]

    def get_queryset(self):
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            return CalendarEvent.objects.filter(project_id=project_pk)
        return CalendarEvent.objects.all()

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs.get('project_pk'))
        serializer.save(project=project)


class WeeklyReportViewSet(viewsets.ModelViewSet):
    serializer_class = WeeklyReportSerializer
    permission_classes = [permissions.IsAuthenticated, IsProjectTeamMember]

    def get_queryset(self):
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            return WeeklyReport.objects.filter(project_id=project_pk)
        return WeeklyReport.objects.all()

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs.get('project_pk'))
        serializer.save(project=project)


class ProjectFileViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectFileSerializer
    permission_classes = [permissions.IsAuthenticated, IsProjectTeamMember]

    def get_queryset(self):
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            return ProjectFile.objects.filter(project_id=project_pk)
        return ProjectFile.objects.all()

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs.get('project_pk'))
        serializer.save(project=project)


class ProjectPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectPaymentSerializer

    def get_queryset(self):
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            return ProjectPayment.objects.filter(project_id=project_pk).order_by('-date')
        return ProjectPayment.objects.all()

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs.get('project_pk'))
        serializer.save(project=project)

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdminUser()]
        return [permissions.IsAuthenticated(), IsProjectTeamMember()]


class ProjectExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectExpenseSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            return ProjectExpense.objects.filter(project_id=project_pk).order_by('-date')
        return ProjectExpense.objects.all()

    def perform_create(self, serializer):
        project = get_object_or_404(Project, pk=self.kwargs.get('project_pk'))
        serializer.save(project=project)


class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(recipient=self.request.user)

    @action(detail=True, methods=['post'])
    def read(self, request, pk=None):
        notif = self.get_object()
        notif.is_read = True
        notif.save()
        return Response({'status': 'read'})

    @action(detail=False, methods=['post'])
    def read_all(self, request):
        self.get_queryset().update(is_read=True)
        return Response({'status': 'all_read'})


class DashboardStatsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role == 'admin' or user.is_superuser:
            total_projects = Project.objects.count()
            active_projects = Project.objects.filter(is_started=True).count()
            inactive_projects = total_projects - active_projects

            total_income = ProjectPayment.objects.aggregate(sum=Sum('amount'))['sum'] or 0
            total_expense = ProjectExpense.objects.aggregate(sum=Sum('amount'))['sum'] or 0
            net_profit = total_income - total_expense

            data = {
                "is_admin": True,
                "project_stats": [
                    {"name": "فعال", "value": active_projects, "color": "#00C49F"},
                    {"name": "غیرفعال", "value": inactive_projects, "color": "#FF8042"},
                ],
                "financial_stats": [
                    {"name": "درآمد", "amount": total_income, "fill": "#82ca9d"},
                    {"name": "هزینه", "amount": total_expense, "fill": "#ff8042"},
                    {"name": "سود", "amount": net_profit, "fill": "#ffc658"},
                ]
            }
            return Response(data)
        else:
            my_projects = Project.objects.filter(client_user=user)
            active_count = my_projects.filter(is_started=True).count()
            total_paid = \
            ProjectPayment.objects.filter(project__in=my_projects, is_paid=True).aggregate(sum=Sum('amount'))[
                'sum'] or 0
            total_contract = my_projects.aggregate(sum=Sum('contract_amount'))['sum'] or 0
            total_due = total_contract - total_paid

            data = {
                "is_admin": False,
                "stats": {
                    "active_projects": active_count,
                    "total_projects": my_projects.count(),
                    "total_paid": total_paid,
                    "total_due": total_due,
                }
            }
            return Response(data)


class ProjectFinancialExportView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get(self, request, project_pk):
        try:
            project = Project.objects.get(pk=project_pk)
        except Project.DoesNotExist:
            return Response({"error": "Project not found"}, status=404)

        payments = ProjectPayment.objects.filter(project=project).order_by('date')
        expenses = ProjectExpense.objects.filter(project=project).order_by('date')

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "خلاصه وضعیت"
        ws_summary.append(["عنوان", "مقدار (تومان)"])

        total_received = payments.aggregate(sum=Sum('amount'))['sum'] or 0
        total_expenses = expenses.aggregate(sum=Sum('amount'))['sum'] or 0
        net_profit = total_received - total_expenses

        ws_summary.append(["نام پروژه", project.project_name])
        ws_summary.append(["مبلغ قرارداد", project.contract_amount])
        ws_summary.append(["کل دریافتی", total_received])
        ws_summary.append(["کل هزینه‌ها", total_expenses])
        ws_summary.append(["سود خالص", net_profit])

        ws_payments = wb.create_sheet("دریافتی‌ها")
        ws_payments.append(["تاریخ", "مبلغ", "بابت"])
        for p in payments:
            ws_payments.append([str(p.date), p.amount, p.description])

        ws_expenses = wb.create_sheet("هزینه‌ها")
        ws_expenses.append(["تاریخ", "مبلغ", "بابت"])
        for e in expenses:
            ws_expenses.append([str(e.date), e.amount, e.description])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Financial_Report_{project.id}.xlsx'
        wb.save(response)
        return response


class SalaryPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = SalaryPaymentSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return SalaryPayment.objects.all().order_by('-payment_date')


class GeneralExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = GeneralExpenseSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return GeneralExpense.objects.all().order_by('-date')


class ScenarioCommentViewSet(viewsets.ModelViewSet):
    serializer_class = ScenarioCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        scenario_id = self.request.query_params.get('scenario')
        if scenario_id:
            return ScenarioComment.objects.filter(scenario_id=scenario_id).order_by('created_at')
        return ScenarioComment.objects.none()

    def perform_create(self, serializer):
        serializer.save(author=self.request.user)


class GlobalCalendarEventViewSet(viewsets.ModelViewSet):
    serializer_class = GlobalCalendarEventSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin' or user.is_superuser:
            return CalendarEvent.objects.all().order_by('event_date')
        if user.role == 'client':
            return CalendarEvent.objects.filter(project__client_user=user).order_by('event_date')
        return CalendarEvent.objects.filter(
            Q(project__writers=user) |
            Q(project__videographers=user) |
            Q(project__editors=user) |
            Q(project__designers=user) |
            Q(project__social_admins=user)
        ).distinct().order_by('event_date')

    def perform_create(self, serializer):
        user = self.request.user
        if user.role != 'admin' and not user.is_superuser:
            pass
        serializer.save()


def log_activity(user, action, model, desc, project=None, request=None):
    ip = None
    if request:
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')

    ActivityLog.objects.create(
        user=user,
        action_type=action,
        model_name=model,
        description=desc,
        project=project,
        ip_address=ip
    )


class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ActivityLogSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return ActivityLog.objects.all().order_by('-created_at')


class ChatRoomViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ChatRoomSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return self.request.user.chat_rooms.all().order_by('-updated_at')


class ChatMessageViewSet(viewsets.ModelViewSet):
    serializer_class = ChatMessageSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        room_id = self.request.query_params.get('room')
        if room_id:
            if self.request.user.chat_rooms.filter(id=room_id).exists():
                return ChatMessage.objects.filter(room_id=room_id).order_by('created_at')
        return ChatMessage.objects.none()

    def perform_create(self, serializer):
        room_id = self.request.data.get('room')
        room = get_object_or_404(ChatRoom, id=room_id)
        if self.request.user not in room.participants.all():
            raise permissions.PermissionDenied("شما عضو این گفتگو نیستید.")
        serializer.save(sender=self.request.user, room=room)
        room.updated_at = timezone.now()
        room.save()


# ✅ ویو اصلاح شده با مدل شیائومی و قابلیت Reasoning
class ProjectAIAnalysisView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, project_pk):
        print(f"🔹 Start AI Analysis for project: {project_pk} using {OPENROUTER_MODEL}")
        try:
            project = get_object_or_404(Project, pk=project_pk)
            reports = WeeklyReport.objects.filter(project=project).order_by('week_number')
            prompt_text = f"نام پروژه: {project.project_name}\nهدف ماهانه: {project.monthly_post_goal} پست\nگزارش‌ها:\n"

            if not reports.exists():
                return Response({"analysis": "هنوز گزارشی ثبت نشده است."}, status=200)

            for r in reports:
                prompt_text += f"- هفته {r.week_number}: {r.report_text}\n"

            prompt_text += "\nلطفاً نقاط قوت، ضعف و برنامه ماه بعد را خلاصه بگو."

            url = "https://openrouter.ai/api/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json",
                "HTTP-Referer": "http://localhost:3000",
                "X-Title": "My CRM",
            }
            # ✅ اضافه کردن تنظیمات Reasoning طبق درخواست
            payload = {
                "model": OPENROUTER_MODEL,
                "messages": [
                    {"role": "system", "content": "You are a helpful marketing assistant."},
                    {"role": "user", "content": prompt_text}
                ],
                "reasoning": {"enabled": True}  # فعال‌سازی استدلال برای این مدل خاص
            }

            print(f"📡 Sending request to OpenRouter...")
            response = requests.post(url, json=payload, headers=headers, timeout=60)

            print(f"📩 OpenRouter Response Status: {response.status_code}")

            if response.ok:
                data = response.json()
                analysis_text = data["choices"][0]["message"]["content"]
                return Response({"analysis": analysis_text})
            else:
                print(f"❌ OpenRouter Error: {response.text}")
                return Response({"analysis": f"خطای هوش مصنوعی ({response.status_code}): {response.text}"},
                                status=response.status_code)

        except Exception as e:
            print(f"❌ Server Error: {str(e)}")
            return Response({"analysis": f"خطای داخلی سرور: {str(e)}"}, status=500)


class PackageViewSet(viewsets.ModelViewSet):
    serializer_class = PackageSerializer
    queryset = Package.objects.all()

    def get_permissions(self):
        if self.action in ['list', 'retrieve']: return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsAdminUser()]


class PaymentMethodViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentMethodSerializer
    queryset = PaymentMethod.objects.all()

    def get_permissions(self):
        if self.action in ['list', 'retrieve']: return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsAdminUser()]


class AgencyInfoViewSet(viewsets.ModelViewSet):
    serializer_class = AgencyInfoSerializer
    queryset = AgencyInfo.objects.all()

    def get_permissions(self):
        if self.action in ['list', 'retrieve']: return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsAdminUser()]


class ExtraServiceViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ExtraService.objects.all()
    serializer_class = ExtraServiceSerializer
    permission_classes = [permissions.IsAuthenticated]


class ServiceRequestViewSet(viewsets.ModelViewSet):
    serializer_class = ServiceRequestSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin' or user.is_superuser:
            return ServiceRequest.objects.all().order_by('-created_at')
        return ServiceRequest.objects.filter(project__client_user=user).order_by('-created_at')

    def perform_create(self, serializer):
        project = serializer.validated_data['project']
        user = self.request.user
        if user.role != 'admin' and project.client_user != user:
            raise permissions.PermissionDenied("شما اجازه ثبت سفارش برای این پروژه را ندارید.")
        serializer.save()


class AgencyFileViewSet(viewsets.ModelViewSet):
    serializer_class = AgencyFileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role == 'admin' or self.request.user.is_superuser:
            return AgencyFile.objects.all().order_by('-uploaded_at')
        return AgencyFile.objects.none()

    def perform_create(self, serializer):
        if self.request.user.role != 'admin' and not self.request.user.is_superuser:
            raise permissions.PermissionDenied("فقط مدیران می‌توانند فایل آژانس آپلود کنند.")
        file = self.request.FILES['file']
        ext = file.name.split('.')[-1].lower()
        f_type = 'document'
        if ext in ['jpg', 'jpeg', 'png', 'gif']:
            f_type = 'image'
        elif ext in ['mp4', 'mov', 'avi']:
            f_type = 'video'
        elif ext in ['pdf']:
            f_type = 'pdf'
        elif ext in ['doc', 'docx', 'xls', 'xlsx']:
            f_type = 'office'
        serializer.save(file_type=f_type)


class TargetAudienceViewSet(viewsets.ModelViewSet):
    serializer_class = TargetAudienceSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

    def get_queryset(self):
        return TargetAudience.objects.all().order_by('-created_at')


# ✅ ویو اصلاح شده تولید سناریو با شیائومی و Reasoning
class ProjectScenarioGenerationView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, project_pk):
        print(f"🔹 Generating Scenario for project: {project_pk} using {OPENROUTER_MODEL}")
        try:
            project = get_object_or_404(Project, pk=project_pk)
            topic = request.data.get('topic', 'آزاد')
            style = request.data.get('style', 'آموزشی')

            prompt_text = (
                f"You are a professional scriptwriter. Project: {project.project_name}. Topic: {topic}. Style: {style}. "
                f"Write a Persian scenario. Output VALID JSON ONLY: {{'title': '', 'summary': '', 'full_script': ''}}"
            )

            url = "https://openrouter.ai/api/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json"
            }
            # ✅ اضافه کردن تنظیمات Reasoning
            payload = {
                "model": OPENROUTER_MODEL,
                "messages": [
                    {"role": "system", "content": "You are a helpful assistant that outputs JSON only."},
                    {"role": "user", "content": prompt_text}
                ],
                "reasoning": {"enabled": True}
            }

            print(f"📡 Sending Scenario Request...")
            response = requests.post(url, json=payload, headers=headers, timeout=60)
            print(f"📩 Response Status: {response.status_code}")

            if response.ok:
                data = response.json()
                ai_content = data["choices"][0]["message"]["content"]

                # تمیزکاری JSON
                if "```json" in ai_content:
                    ai_content = ai_content.split("```json")[1].split("```")[0].strip()
                elif "```" in ai_content:
                    ai_content = ai_content.split("```")[1].split("```")[0].strip()

                try:
                    return Response(json.loads(ai_content))
                except json.JSONDecodeError:
                    return Response({
                        "title": f"سناریو درباره {topic}",
                        "summary": "تولید شده توسط هوش مصنوعی (نیاز به بازبینی)",
                        "full_script": ai_content
                    })
            else:
                print(f"❌ Error: {response.text}")
                return Response({"error": f"خطا در ارتباط با هوش مصنوعی: {response.text}"}, status=500)

        except Exception as e:
            print(f"❌ Server Exception: {e}")
            return Response({"error": str(e)}, status=500)


# ✅ ویو اصلاح شده تحلیل مخاطب با شیائومی و Reasoning
class TargetAudienceAIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        topic = request.data.get('topic')
        if not topic:
            return Response({"error": "موضوع را وارد کنید"}, status=400)

        prompt_text = (
            f"Act as a marketing strategist. Analyze the target audience for: '{topic}'. "
            f"Output Persian JSON keys: title, icon_name, job_title, age_range, gender, income_level, pain_points, goals, our_solution"
        )

        try:
            url = "https://openrouter.ai/api/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                "Content-Type": "application/json"
            }
            # ✅ اضافه کردن تنظیمات Reasoning
            payload = {
                "model": OPENROUTER_MODEL,
                "messages": [
                    {"role": "system", "content": "You are a marketing expert who speaks Persian and outputs JSON."},
                    {"role": "user", "content": prompt_text}
                ],
                "reasoning": {"enabled": True}
            }

            print(f"📡 Sending Target Audience Request...")
            response = requests.post(url, json=payload, headers=headers, timeout=60)
            print(f"📩 Response Status: {response.status_code}")

            if response.ok:
                data = response.json()
                content = data["choices"][0]["message"]["content"]
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()
                return Response(json.loads(content))
            else:
                print(f"❌ Error: {response.text}")
                return Response({"error": f"AI Error: {response.text}"}, status=500)
        except Exception as e:
            print(f"❌ Server Exception: {e}")
            return Response({"error": str(e)}, status=500)


class TaskViewSet(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['title', 'description']
    ordering_fields = ['due_date', 'priority', 'created_at']
    filterset_fields = ['status', 'priority', 'assigned_to']

    def get_queryset(self):
        project_id = self.kwargs.get('project_pk')
        if project_id:
            return Task.objects.filter(project_id=project_id)
        return Task.objects.all()

    def perform_create(self, serializer):
        project_id = self.kwargs.get('project_pk')
        if project_id:
            project = get_object_or_404(Project, pk=project_id)
            serializer.save(project=project)
        else:
            serializer.save()


class GlobalSearchView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        query = request.query_params.get('q', '').strip()
        if not query:
            return Response([])

        results = []
        user = request.user
        is_admin = user.role == 'admin' or user.is_superuser

        projects = Project.objects.filter(project_name__icontains=query, is_deleted=False)
        if not is_admin:
            projects = projects.filter(client_user=user)
        for p in projects[:5]:
            results.append(
                {"type": "project", "title": p.project_name, "subtitle": "پروژه", "link": f"/project/{p.id}"})

        if is_admin:
            users = User.objects.filter(Q(username__icontains=query) | Q(full_name__icontains=query), is_deleted=False)[
                    :5]
            for u in users:
                results.append({"type": "user", "title": u.full_name or u.username, "subtitle": u.get_role_display(),
                                "link": "/users"})

        scenarios = Scenario.objects.filter(title__icontains=query)[:5]
        for s in scenarios:
            if is_admin or s.project.client_user == user:
                results.append({"type": "scenario", "title": s.title, "subtitle": f"سناریو در {s.project.project_name}",
                                "link": f"/project/{s.project.id}?tab=scenarios"})

        if is_admin:
            payments = ProjectPayment.objects.filter(description__icontains=query)[:5]
            for pay in payments:
                results.append(
                    {"type": "invoice", "title": f"فاکتور: {pay.description}", "subtitle": f"{pay.amount:,} تومان",
                     "link": f"/project/{pay.project.id}?tab=financials"})

        return Response(results)


class FileCommentViewSet(viewsets.ModelViewSet):
    serializer_class = FileCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.detail:
            return FileComment.objects.all()
        file_id = self.request.query_params.get('file_id')
        if file_id:
            return FileComment.objects.filter(file_id=file_id).order_by('created_at')
        return FileComment.objects.none()

    def perform_create(self, serializer):
        serializer.save(author=self.request.user)


class StickyNoteViewSet(viewsets.ModelViewSet):
    serializer_class = StickyNoteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return StickyNote.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class WorkflowRuleViewSet(viewsets.ModelViewSet):
    serializer_class = WorkflowRuleSerializer
    permission_classes = [permissions.IsAdminUser]
    queryset = WorkflowRule.objects.all()


class TimeLogViewSet(viewsets.ModelViewSet):
    serializer_class = TimeLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.role == 'admin' or self.request.user.is_superuser:
            return TimeLog.objects.all().order_by('-start_time')
        return TimeLog.objects.filter(user=self.request.user).order_by('-start_time')

    @action(detail=False, methods=['post'])
    def start_timer(self, request):
        task_id = request.data.get('task_id')
        if not task_id: return Response({'error': 'Task ID is required'}, status=400)
        active_log = TimeLog.objects.filter(user=request.user, end_time__isnull=True).first()
        if active_log: return Response({'error': 'یک تایمر فعال دارید.'}, status=400)
        task = Task.objects.get(pk=task_id)
        log = TimeLog.objects.create(user=request.user, task=task, start_time=timezone.now())
        return Response(TimeLogSerializer(log).data)

    @action(detail=False, methods=['post'])
    def stop_timer(self, request):
        active_log = TimeLog.objects.filter(user=request.user, end_time__isnull=True).first()
        if not active_log: return Response({'error': 'هیچ تایمر فعالی یافت نشد.'}, status=404)
        active_log.end_time = timezone.now()
        active_log.description = request.data.get('description', '')
        active_log.save()
        return Response(TimeLogSerializer(active_log).data)

    @action(detail=False, methods=['get'])
    def current(self, request):
        active_log = TimeLog.objects.filter(user=request.user, end_time__isnull=True).first()
        if active_log: return Response(TimeLogSerializer(active_log).data)
        return Response(None)


class SharedLinkViewSet(viewsets.ModelViewSet):
    serializer_class = SharedLinkSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return SharedLink.objects.all()

    @action(detail=False, methods=['post'])
    def generate(self, request):
        file_id = request.data.get('file_id')
        try:
            file_obj = ProjectFile.objects.get(id=file_id)
            link, created = SharedLink.objects.get_or_create(file=file_obj)
            return Response({'link_id': link.id, 'url': f"/share/{link.id}"})
        except ProjectFile.DoesNotExist:
            return Response({'error': 'File not found'}, status=404)


class PublicFileView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, token):
        try:
            share = SharedLink.objects.get(id=token, is_active=True)
            serializer = SharedLinkSerializer(share)
            return Response(serializer.data)
        except SharedLink.DoesNotExist:
            return Response({'error': 'Invalid or expired link'}, status=404)

    def post(self, request, token):
        try:
            share = SharedLink.objects.get(id=token, is_active=True)
            data = request.data
            comment = FileComment.objects.create(
                file=share.file,
                author=None,
                guest_name=data.get('guest_name', 'مهمان'),
                text=data.get('text'),
                x_position=data.get('x_position'),
                y_position=data.get('y_position')
            )
            return Response(FileCommentSerializer(comment).data)
        except SharedLink.DoesNotExist:
            return Response({'error': 'Invalid link'}, status=404)


class DashboardConfigViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        config, created = DashboardConfig.objects.get_or_create(user=request.user)
        return Response({'active_widgets': config.active_widgets})

    def create(self, request):
        config, created = DashboardConfig.objects.get_or_create(user=request.user)
        config.active_widgets = request.data.get('widgets', [])
        config.save()
        return Response({'status': 'saved', 'active_widgets': config.active_widgets})


class LeadViewSet(viewsets.ModelViewSet):
    queryset = Lead.objects.all().order_by('-created_at')
    serializer_class = LeadSerializer
    permission_classes = [permissions.IsAuthenticated]